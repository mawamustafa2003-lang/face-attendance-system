import cv2
import os
import numpy as np
import tkinter as tk
from tkinter import ttk, messagebox
from PIL import Image, ImageTk
import datetime
import openpyxl

# ================= SETTINGS =================
DATASET_DIR = "dataset"
ATTENDANCE_FILE = "attendance.xlsx"
os.makedirs(DATASET_DIR, exist_ok=True)

# حذف التقرير القديم عند كل تشغيل (اختياري)
if os.path.exists(ATTENDANCE_FILE):
    os.remove(ATTENDANCE_FILE)

face_detector = cv2.CascadeClassifier(
    cv2.data.haarcascades + "haarcascade_frontalface_default.xml"
)

recognizer = cv2.face.LBPHFaceRecognizer_create()

students = {}
camera_running = False
cap = None
model_trained = False  # <-- مهم: لمعرفة هل تم تدريب الموديل

CONFIDENCE_THRESHOLD = 55


# ================= TRAIN MODEL =================
def train_model():
    global model_trained
    faces = []
    labels = []

    students.clear()
    label_id = 0

    for folder in os.listdir(DATASET_DIR):
        path = os.path.join(DATASET_DIR, folder)

        if not os.path.isdir(path):
            continue

        students[label_id] = folder

        for img_name in os.listdir(path):
            img_path = os.path.join(path, img_name)

            img = cv2.imread(img_path, cv2.IMREAD_GRAYSCALE)

            if img is None:
                continue

            img = cv2.equalizeHist(img)

            faces.append(img)
            labels.append(label_id)

        label_id += 1

    if len(faces) > 0:
        recognizer.train(faces, np.array(labels))
        model_trained = True
        print("Model trained successfully")
    else:
        model_trained = False
        print("Dataset is empty - recognition disabled")


# ================= REGISTER =================
def register_student():
    name = entry_name.get().strip()
    dept = entry_dept.get().strip()
    subject = entry_subject.get().strip()

    if not name or not dept or not subject:
        messagebox.showerror("Error", "Fill all fields")
        return

    folder_name = f"{name}_{dept}_{subject}"
    path = os.path.join(DATASET_DIR, folder_name)
    os.makedirs(path, exist_ok=True)

    cap_temp = cv2.VideoCapture(0)
    count = 0

    while count < 25:
        ret, frame = cap_temp.read()
        gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)

        faces = face_detector.detectMultiScale(gray, 1.3, 5)

        for (x, y, w, h) in faces:
            face = gray[y:y + h, x:x + w]
            face = cv2.resize(face, (200, 200))
            face = cv2.equalizeHist(face)

            cv2.imwrite(os.path.join(path, f"{count}.jpg"), face)
            count += 1

            cv2.rectangle(frame, (x, y), (x + w, y + h), (0, 255, 0), 2)

        cv2.imshow("Registering...", frame)

        if cv2.waitKey(1) == 27:
            break

    cap_temp.release()
    cv2.destroyAllWindows()

    train_model()

    messagebox.showinfo("Success", "Student Registered")


# ================= SAVE EXCEL =================
def save_attendance(folder):
    name, dept, subject = folder.split("_")

    today = datetime.date.today().isoformat()
    time = datetime.datetime.now().strftime("%H:%M:%S")

    if os.path.exists(ATTENDANCE_FILE):
        wb = openpyxl.load_workbook(ATTENDANCE_FILE)
        ws = wb.active
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Name", "Department", "Subject", "Date", "Time"])

    for row in ws.iter_rows(values_only=True):
        if row[0] == name and row[3] == today:
            return False

    ws.append([name, dept, subject, today, time])
    wb.save(ATTENDANCE_FILE)

    return True


# ================= CAMERA LOOP =================
def update_camera():
    global cap, camera_running

    if not camera_running:
        return

    ret, frame = cap.read()
    gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)

    faces = face_detector.detectMultiScale(gray, 1.3, 5)

    for (x, y, w, h) in faces:

        face = gray[y:y + h, x:x + w]
        face = cv2.resize(face, (200, 200))
        face = cv2.equalizeHist(face)

        name = "Unknown"
        color = (0, 0, 255)

        # إذا لم يتم تدريب الموديل لا نحاول التعرف
        if model_trained:

            label, confidence = recognizer.predict(face)

            if confidence < CONFIDENCE_THRESHOLD and label in students:

                folder = students[label]
                saved = save_attendance(folder)

                name = folder.split("_")[0]
                color = (0, 255, 0)

                if saved:
                    status_label.config(text=f"✔ {name} Marked", fg="green")
                    refresh_table()
                else:
                    status_label.config(text=f"{name} Already Marked Today", fg="orange")

        cv2.rectangle(frame, (x, y), (x + w, y + h), color, 2)
        cv2.putText(frame, name, (x, y - 10),
                    cv2.FONT_HERSHEY_SIMPLEX, 0.8, color, 2)

    img = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
    img = ImageTk.PhotoImage(Image.fromarray(img))

    camera_label.imgtk = img
    camera_label.configure(image=img)

    root.after(10, update_camera)


# ================= CAMERA CONTROL =================
def start_camera():
    global cap, camera_running

    if not camera_running:
        cap = cv2.VideoCapture(0)
        camera_running = True
        update_camera()


def stop_camera():
    global camera_running, cap

    camera_running = False

    if cap:
        cap.release()


# ================= TABLE =================
def refresh_table():

    for row in table.get_children():
        table.delete(row)

    if not os.path.exists(ATTENDANCE_FILE):
        return

    wb = openpyxl.load_workbook(ATTENDANCE_FILE)
    ws = wb.active

    for row in ws.iter_rows(min_row=2, values_only=True):
        table.insert("", "end", values=row)


# ================= GUI =================
root = tk.Tk()
root.title("Smart Attendance Pro")
root.geometry("1000x600")
root.configure(bg="#0f172a")

left_frame = tk.Frame(root, bg="#1e293b")
left_frame.pack(side="left", fill="both", expand=True)

right_frame = tk.Frame(root, bg="#111827")
right_frame.pack(side="right", fill="both")

camera_label = tk.Label(left_frame)
camera_label.pack()

status_label = tk.Label(
    left_frame,
    text="",
    bg="#1e293b",
    font=("Arial", 14, "bold"),
    fg="white"
)
status_label.pack(pady=10)

form_frame = tk.Frame(right_frame, bg="#111827", pady=20)
form_frame.pack()

entry_name = tk.Entry(form_frame, width=25, font=("Arial", 14))
entry_dept = tk.Entry(form_frame, width=25, font=("Arial", 14))
entry_subject = tk.Entry(form_frame, width=25, font=("Arial", 14))

tk.Label(form_frame, text="Name", bg="#111827", fg="white",
         font=("Arial", 14)).pack()
entry_name.pack(pady=5)

tk.Label(form_frame, text="Department", bg="#111827", fg="white",
         font=("Arial", 14)).pack()
entry_dept.pack(pady=5)

tk.Label(form_frame, text="Subject", bg="#111827", fg="white",
         font=("Arial", 14)).pack()
entry_subject.pack(pady=5)

button_style = {
    "font": ("Arial", 14),
    "width": 20,
    "height": 1
}

tk.Button(form_frame, text="Register Student",
          command=register_student,
          bg="#22c55e", fg="white",
          **button_style).pack(pady=5)

tk.Button(form_frame, text="Start Camera",
          command=start_camera,
          bg="#2563eb", fg="white",
          **button_style).pack(pady=5)

tk.Button(form_frame, text="Stop Camera",
          command=stop_camera,
          bg="#ef4444", fg="white",
          **button_style).pack(pady=5)

tk.Button(form_frame, text="Export Excel",
          command=lambda: messagebox.showinfo(
              "Saved", f"Report saved as {ATTENDANCE_FILE}"
          ),
          bg="#f59e0b", fg="white",
          **button_style).pack(pady=5)

table = ttk.Treeview(right_frame,
                     columns=("Name", "Department", "Subject", "Date", "Time"),
                     show="headings")

for col in ("Name", "Department", "Subject", "Date", "Time"):
    table.heading(col, text=col)
    table.column(col, width=120)

table.pack(fill="both", expand=True, pady=10)

train_model()
refresh_table()

root.mainloop()